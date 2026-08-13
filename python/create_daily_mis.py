import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ============================================================
# 1. PROJECT PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent

PROCESSED_DATA_DIR = BASE_DIR / "data" / "processed"
REPORTS_DIR = BASE_DIR / "reports"

REPORTS_DIR.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# 2. FILE PATHS
# ============================================================

OPERATIONS_FILE = (
    PROCESSED_DATA_DIR /
    "operations_processed.csv"
)

TEAM_FILE = (
    PROCESSED_DATA_DIR /
    "team_performance.csv"
)

SHIFT_FILE = (
    PROCESSED_DATA_DIR /
    "shift_performance.csv"
)

DAILY_FILE = (
    PROCESSED_DATA_DIR /
    "daily_performance.csv"
)

EMPLOYEE_FILE = (
    PROCESSED_DATA_DIR /
    "employee_performance.csv"
)

REPORT_FILE = (
    REPORTS_DIR /
    "Daily_Operations_MIS.xlsx"
)

BACKLOG_FILE = (
    PROCESSED_DATA_DIR /
    "daily_backlog.csv"
)


# ============================================================
# 3. BUSINESS CONSTANTS
# ============================================================

PRODUCTIVITY_BENCHMARK = 15
SLA_TARGET = 95
HIGH_ERROR_THRESHOLD = 5


# ============================================================
# 4. LOAD DATA
# ============================================================

print("\n" + "=" * 60)
print("LOADING MANAGEMENT DATA")
print("=" * 60)

operations = pd.read_csv(OPERATIONS_FILE)
team = pd.read_csv(TEAM_FILE)
shift = pd.read_csv(SHIFT_FILE)
daily = pd.read_csv(DAILY_FILE)
employees = pd.read_csv(EMPLOYEE_FILE)

operations["Date"] = pd.to_datetime(
    operations["Date"]
)

daily["Date"] = pd.to_datetime(
    daily["Date"]
)

print(f"\nOperations rows: {len(operations)}")
print(f"Daily rows: {len(daily)}")
print(f"Teams: {len(team)}")
print(f"Shifts: {len(shift)}")
print(f"Employees: {len(employees)}")


# ============================================================
# 5. BUILD CUMULATIVE BACKLOG
# ============================================================

print("\n" + "=" * 60)
print("CALCULATING CUMULATIVE BACKLOG")
print("=" * 60)

daily_backlog = (
    daily[
        [
            "Date",
            "Received",
            "Processed",
            "Pending"
        ]
    ]
    .sort_values("Date")
    .reset_index(drop=True)
)

daily_backlog["Opening Backlog"] = 0
daily_backlog["Closing Backlog"] = 0

opening_backlog = 0

for index in range(len(daily_backlog)):

    received = daily_backlog.loc[
        index,
        "Received"
    ]

    processed = daily_backlog.loc[
        index,
        "Processed"
    ]

    daily_backlog.loc[
        index,
        "Opening Backlog"
    ] = opening_backlog

    closing_backlog = (
        opening_backlog
        + received
        - processed
    )

    daily_backlog.loc[
        index,
        "Closing Backlog"
    ] = closing_backlog

    opening_backlog = closing_backlog


daily_backlog["Backlog Change"] = (
    daily_backlog["Closing Backlog"]
    -
    daily_backlog["Opening Backlog"]
)


# ============================================================
# 6. VALIDATE BACKLOG
# ============================================================

negative_backlog = (
    daily_backlog["Closing Backlog"] < 0
).sum()

print(
    f"Negative closing backlog days: "
    f"{negative_backlog}"
)

print(
    f"Final closing backlog: "
    f"{daily_backlog['Closing Backlog'].iloc[-1]}"
)

daily_backlog.to_csv(
    BACKLOG_FILE,
    index=False
)


# ============================================================
# 7. SELECT LATEST MIS DATE
# ============================================================

latest_date = daily["Date"].max()

print(
    f"\nMIS Date: "
    f"{latest_date.strftime('%d-%b-%Y')}"
)

latest_daily = daily[
    daily["Date"] == latest_date
].iloc[0]

latest_backlog = daily_backlog[
    daily_backlog["Date"] == latest_date
].iloc[0]


# ============================================================
# 8. CALCULATE DAILY KPIs
# ============================================================

received = latest_daily["Received"]
processed = latest_daily["Processed"]
pending = latest_daily["Pending"]
errors = latest_daily["Errors"]
rework = latest_daily["Rework"]

working_hours = latest_daily[
    "Working_Hours"
]

sla_achieved = latest_daily[
    "SLA_Achieved"
]

attendance = latest_daily[
    "Attendance"
]


productivity = (
    processed /
    working_hours
    if working_hours > 0
    else 0
)

productivity_pct = (
    productivity /
    PRODUCTIVITY_BENCHMARK
    * 100
)

accuracy = (
    (
        processed - errors
    )
    /
    processed
    * 100
    if processed > 0
    else 0
)

error_rate = (
    errors /
    processed
    * 100
    if processed > 0
    else 0
)

rework_rate = (
    rework /
    processed
    * 100
    if processed > 0
    else 0
)

sla_pct = (
    sla_achieved /
    processed
    * 100
    if processed > 0
    else 0
)


# ============================================================
# 9. FIND TEAM ISSUES
# ============================================================

highest_backlog_team = team.loc[
    team["Backlog"].idxmax()
]

lowest_sla_team = team.loc[
    team["SLA %"].idxmin()
]

highest_error_team = team.loc[
    team["Error Rate %"].idxmax()
]

lowest_productivity_team = team.loc[
    team["Productivity %"].idxmin()
]


# ============================================================
# 10. FIND SHIFT ISSUES
# ============================================================

lowest_sla_shift = shift.loc[
    shift["SLA %"].idxmin()
]

lowest_productivity_shift = shift.loc[
    shift["Productivity %"].idxmin()
]


# ============================================================
# 11. FIND HIGH-ERROR EMPLOYEES
# ============================================================

high_error_employees = employees[
    employees["Error Rate %"]
    > HIGH_ERROR_THRESHOLD
].copy()


# ============================================================
# 12. GENERATE KEY ISSUES
# ============================================================

key_issues = []

if highest_backlog_team["Backlog"] > 0:

    key_issues.append(
        f"Backlog is highest in "
        f"{highest_backlog_team['Team']} "
        f"at {highest_backlog_team['Backlog']:,.0f} records."
    )


if sla_pct < SLA_TARGET:

    key_issues.append(
        f"Overall SLA achievement is "
        f"{sla_pct:.1f}%, below the "
        f"{SLA_TARGET}% target."
    )


key_issues.append(
    f"{lowest_sla_shift['Shift']} shift has the "
    f"lowest SLA at "
    f"{lowest_sla_shift['SLA %']:.1f}%."
)


key_issues.append(
    f"{highest_error_team['Team']} has the "
    f"highest error rate at "
    f"{highest_error_team['Error Rate %']:.1f}%."
)


if len(high_error_employees) > 0:

    key_issues.append(
        f"{len(high_error_employees)} employees "
        f"exceed the {HIGH_ERROR_THRESHOLD}% "
        f"error-rate threshold."
    )


# ============================================================
# 13. GENERATE RECOMMENDED ACTIONS
# ============================================================

recommended_actions = []

recommended_actions.append(
    f"Review and redistribute pending workload "
    f"from {highest_backlog_team['Team']}."
)

recommended_actions.append(
    f"Investigate SLA breaches in the "
    f"{lowest_sla_shift['Shift']} shift."
)

recommended_actions.append(
    f"Conduct quality review for "
    f"{highest_error_team['Team']}."
)

recommended_actions.append(
    "Review high-error employee cases and "
    "identify recurring error patterns."
)

recommended_actions.append(
    f"Compare best-performing practices from "
    f"higher-productivity teams with "
    f"{lowest_productivity_team['Team']}."
)


# ============================================================
# 14. CREATE EXCEL WORKBOOK
# ============================================================

print("\n" + "=" * 60)
print("CREATING DAILY MIS REPORT")
print("=" * 60)

workbook = Workbook()


# ============================================================
# 15. DEFINE STYLES
# ============================================================

title_font = Font(
    bold=True,
    size=18
)

section_font = Font(
    bold=True,
    size=13
)

header_font = Font(
    bold=True
)

white_font = Font(
    bold=True,
    color="FFFFFF"
)

kpi_value_font = Font(
    bold=True,
    size=16
)

header_fill = PatternFill(
    "solid",
    fgColor="1F4E78"
)

section_fill = PatternFill(
    "solid",
    fgColor="D9EAF7"
)

kpi_fill = PatternFill(
    "solid",
    fgColor="EAF2F8"
)

good_fill = PatternFill(
    "solid",
    fgColor="C6EFCE"
)

warning_fill = PatternFill(
    "solid",
    fgColor="FFEB9C"
)

bad_fill = PatternFill(
    "solid",
    fgColor="FFC7CE"
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# ============================================================
# 16. HELPER FUNCTIONS
# ============================================================

def style_table_sheet(worksheet):

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:

        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = thin_border

    for row_cells in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row_cells:

            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center"
            )

    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            35
        )

    worksheet.row_dimensions[1].height = 24


def add_kpi_card(
    worksheet,
    label,
    value,
    number_format,
    row,
    column
):

    # Label row
    label_cell = worksheet.cell(
        row=row,
        column=column,
        value=label
    )

    label_cell.font = header_font
    label_cell.fill = section_fill
    label_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    label_cell.border = thin_border

    # Value row
    value_cell = worksheet.cell(
        row=row + 1,
        column=column,
        value=value
    )

    value_cell.font = kpi_value_font
    value_cell.fill = kpi_fill
    value_cell.number_format = number_format
    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    value_cell.border = thin_border

    # Second column of the card
    label_cell_2 = worksheet.cell(
        row=row,
        column=column + 1
    )

    value_cell_2 = worksheet.cell(
        row=row + 1,
        column=column + 1
    )

    label_cell_2.fill = section_fill
    label_cell_2.border = thin_border

    value_cell_2.fill = kpi_fill
    value_cell_2.border = thin_border

    worksheet.merge_cells(
        start_row=row,
        start_column=column,
        end_row=row,
        end_column=column + 1
    )

    worksheet.merge_cells(
        start_row=row + 1,
        start_column=column,
        end_row=row + 1,
        end_column=column + 1
    )


# ============================================================
# 17. EXECUTIVE MIS SHEET
# ============================================================

ws = workbook.active
ws.title = "Daily MIS"

ws.merge_cells("A1:H1")

ws["A1"] = "OPERATIONS DAILY MIS"

ws["A1"].font = title_font

ws["A1"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

ws.merge_cells("A2:H2")

ws["A2"] = (
    f"Date: "
    f"{latest_date.strftime('%d-%b-%Y')}"
)

ws["A2"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

ws["A2"].font = Font(
    italic=True,
    size=11
)


# ============================================================
# 18. KPI CARDS
# ============================================================

add_kpi_card(
    ws,
    "TOTAL RECEIVED",
    received,
    "#,##0",
    4,
    1
)

add_kpi_card(
    ws,
    "TOTAL PROCESSED",
    processed,
    "#,##0",
    4,
    3
)

add_kpi_card(
    ws,
    "PENDING",
    pending,
    "#,##0",
    4,
    5
)

add_kpi_card(
    ws,
    "CLOSING BACKLOG",
    latest_backlog["Closing Backlog"],
    "#,##0",
    4,
    7
)

add_kpi_card(
    ws,
    "SLA ACHIEVEMENT",
    sla_pct / 100,
    "0.00%",
    7,
    1
)

add_kpi_card(
    ws,
    "QUALITY",
    accuracy / 100,
    "0.00%",
    7,
    3
)

add_kpi_card(
    ws,
    "PRODUCTIVITY",
    productivity_pct / 100,
    "0.00%",
    7,
    5
)

add_kpi_card(
    ws,
    "ATTENDANCE",
    attendance / 100,
    "0.00%",
    7,
    7
)


# ============================================================
# 19. KPI CONDITIONAL FORMATTING
# ============================================================

# SLA
ws.conditional_formatting.add(
    "A8",
    CellIsRule(
        operator="lessThan",
        formula=["0.95"],
        fill=bad_fill
    )
)

ws.conditional_formatting.add(
    "A8",
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["0.95"],
        fill=good_fill
    )
)

# Quality
ws.conditional_formatting.add(
    "C8",
    CellIsRule(
        operator="lessThan",
        formula=["0.95"],
        fill=bad_fill
    )
)

ws.conditional_formatting.add(
    "C8",
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["0.95"],
        fill=good_fill
    )
)

# Productivity
ws.conditional_formatting.add(
    "E8",
    CellIsRule(
        operator="lessThan",
        formula=["0.75"],
        fill=bad_fill
    )
)

ws.conditional_formatting.add(
    "E8",
    CellIsRule(
        operator="between",
        formula=["0.75", "0.999999"],
        fill=warning_fill
    )
)

ws.conditional_formatting.add(
    "E8",
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["1"],
        fill=good_fill
    )
)

# Attendance
ws.conditional_formatting.add(
    "G8",
    CellIsRule(
        operator="lessThan",
        formula=["0.95"],
        fill=warning_fill
    )
)

ws.conditional_formatting.add(
    "G8",
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["0.95"],
        fill=good_fill
    )
)


# ============================================================
# 20. BACKLOG STATUS
# ============================================================

ws.merge_cells("A11:H11")

ws["A11"] = "BACKLOG STATUS"

ws["A11"].font = section_font

ws["A11"].fill = section_fill

ws["A11"].alignment = Alignment(
    horizontal="center"
)

backlog_headers = [
    "Opening Backlog",
    "Received",
    "Processed",
    "Closing Backlog",
    "Backlog Change"
]

backlog_values = [
    latest_backlog["Opening Backlog"],
    latest_backlog["Received"],
    latest_backlog["Processed"],
    latest_backlog["Closing Backlog"],
    latest_backlog["Backlog Change"]
]

for index, header in enumerate(
    backlog_headers,
    start=1
):

    cell = ws.cell(
        row=12,
        column=index,
        value=header
    )

    cell.font = white_font
    cell.fill = header_fill
    cell.alignment = Alignment(
        horizontal="center"
    )
    cell.border = thin_border


for index, value in enumerate(
    backlog_values,
    start=1
):

    cell = ws.cell(
        row=13,
        column=index,
        value=value
    )

    cell.number_format = "#,##0"

    cell.alignment = Alignment(
        horizontal="center"
    )

    cell.border = thin_border

    cell.fill = kpi_fill


# Positive backlog change = warning
ws.conditional_formatting.add(
    "E13",
    CellIsRule(
        operator="greaterThan",
        formula=["0"],
        fill=warning_fill
    )
)

# Zero/negative change = good
ws.conditional_formatting.add(
    "E13",
    CellIsRule(
        operator="lessThanOrEqual",
        formula=["0"],
        fill=good_fill
    )
)


# ============================================================
# 21. KEY ISSUES
# ============================================================

ws.merge_cells("A16:H16")

ws["A16"] = "KEY ISSUES"

ws["A16"].font = section_font

ws["A16"].fill = section_fill

row = 17

for issue in key_issues:

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=8
    )

    cell = ws.cell(
        row=row,
        column=1,
        value="• " + issue
    )

    cell.alignment = Alignment(
        wrap_text=True,
        vertical="center"
    )

    row += 1


# ============================================================
# 22. RECOMMENDED ACTIONS
# ============================================================

action_start = row + 1

ws.merge_cells(
    start_row=action_start,
    start_column=1,
    end_row=action_start,
    end_column=8
)

ws.cell(
    row=action_start,
    column=1,
    value="RECOMMENDED ACTIONS"
)

ws.cell(
    row=action_start,
    column=1
).font = section_font

ws.cell(
    row=action_start,
    column=1
).fill = section_fill

row = action_start + 1

for action in recommended_actions:

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=8
    )

    cell = ws.cell(
        row=row,
        column=1,
        value="• " + action
    )

    cell.alignment = Alignment(
        wrap_text=True,
        vertical="center"
    )

    row += 1


# ============================================================
# 23. SUPPORTING SHEET HELPER
# ============================================================

def create_data_sheet(
    workbook,
    sheet_name,
    dataframe
):

    worksheet = workbook.create_sheet(
        sheet_name
    )

    worksheet.append(
        list(dataframe.columns)
    )

    for record in dataframe.itertuples(
        index=False,
        name=None
    ):

        worksheet.append(
            list(record)
        )

    style_table_sheet(
        worksheet
    )

    return worksheet


# ============================================================
# 24. SUPPORTING SHEETS
# ============================================================

team_ws = create_data_sheet(
    workbook,
    "Team Performance",
    team
)

shift_ws = create_data_sheet(
    workbook,
    "Shift Performance",
    shift
)

backlog_ws = create_data_sheet(
    workbook,
    "Daily Backlog",
    daily_backlog
)

employee_ws = create_data_sheet(
    workbook,
    "Employee Performance",
    employees
)


# ============================================================
# 25. DAILY MIS FORMATTING
# ============================================================

ws.sheet_view.showGridLines = False

ws.freeze_panes = "A4"

for column in [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H"
]:

    ws.column_dimensions[column].width = 18


ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 22

for row_number in [
    4,
    7,
    11,
    16
]:

    ws.row_dimensions[
        row_number
    ].height = 24


for row_number in [
    5,
    8
]:

    ws.row_dimensions[
        row_number
    ].height = 30


for row_number in range(
    17,
    ws.max_row + 1
):

    ws.row_dimensions[
        row_number
    ].height = 26


# Apply borders/alignment to executive sheet
for row_cells in ws.iter_rows():

    for cell in row_cells:

        if cell.value is not None:

            cell.border = thin_border

            if cell.alignment is None:

                cell.alignment = Alignment(
                    vertical="center"
                )


# Print settings
ws.page_setup.orientation = "landscape"

ws.page_setup.fitToWidth = 1

ws.page_setup.fitToHeight = 0

ws.sheet_properties.pageSetUpPr.fitToPage = True

ws.print_area = (
    f"A1:H{ws.max_row}"
)


# ============================================================
# 26. SAVE WORKBOOK
# ============================================================

workbook.save(
    REPORT_FILE
)


# ============================================================
# 27. FINAL OUTPUT
# ============================================================

print("\n" + "=" * 60)
print("DAILY MIS REPORT CREATED SUCCESSFULLY")
print("=" * 60)

print(
    "\nReport created:"
)

print(
    REPORT_FILE
)

print(
    f"\nMIS Date: "
    f"{latest_date.strftime('%d-%b-%Y')}"
)

print(
    f"Closing Backlog: "
    f"{latest_backlog['Closing Backlog']:,.0f}"
)

print(
    f"SLA: "
    f"{sla_pct:.2f}%"
)

print(
    f"Quality: "
    f"{accuracy:.2f}%"
)

print(
    f"Productivity: "
    f"{productivity_pct:.2f}%"
)

print(
    f"Attendance: "
    f"{attendance:.2f}%"
)